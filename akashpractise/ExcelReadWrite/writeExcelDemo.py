from openpyxl import Workbook
from openpyxl.styles import PatternFill
'''Creat a Workbook class as a wb '''
wb= Workbook()

'''know the active sheet'''
sheets = wb.active.title
print(sheets)
print(wb.sheetnames)

'''change the sheet name'''
wb["Sheet"].title = "Report of Automation"
sh1 = wb.active
print(sh1)

'''add value in sheet '''
sh1["A1"].value ="Name"
sh1["B1"].value ="Status"
sh1["C1"].value ="Duration"

sh1["A2"].value ="Python"
sh1["B2"].value ="Active"
sh1["B2"].fill =PatternFill("solid",fgColor="FF5733")
sh1["C2"].value ="10"

sh1["A3"].value ="Java"
sh1["B3"].value ="Deactive"
sh1["B3"].fill =PatternFill("solid",fgColor="0FE800")
sh1["C3"].value ="15"

sh1["A4"].value ="C++"
sh1["B4"].value ="Active"
sh1["B4"].fill =PatternFill("solid",fgColor="FF5733")
sh1["C4"].value ="20"

'''Save file with name'''
# wb.save("AutomationFile.xlsx")

'''save file as perticular loction'''
wb.save("C:\\Users\\AKASH\\OneDrive\\Desktop\\AutomationReport.xlsx")

'''video link
https://www.youtube.com/watch?v=nsKNPHJ9iPc'''