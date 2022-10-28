import openpyxl
wb = openpyxl.load_workbook("C:\\Users\AKASH\\OneDrive\\Desktop\\Employee.xlsx")
print(type(wb))
sheets = wb.sheetnames
print(sheets)
'''for current active sheet'''
print(wb.active.title)

'''select perticular sheet'''
sh1= wb["Employe Data"]
print(type(sh1))

'''way 1:   cell data, sheet>cell>Get Data'''
data=wb["Employe Data"]["B2"].value
# print(data)

'''way2 for read: .cell(Row, column)'''
print(sh1.cell(4,1).value)

sh2 = wb["Job Role"]
print(sh2.cell(4,3).value, "job role")

'''Way 3 for read: '''
print(sh2.cell(row=2,column=3).value)

print(wb.get_sheet_by_name("Employe Data").cell(row=4,column=1).value)