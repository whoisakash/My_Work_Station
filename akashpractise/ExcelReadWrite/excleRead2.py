import openpyxl
wb = openpyxl.load_workbook("C:\\Users\AKASH\\OneDrive\\Desktop\\Employee.xlsx")
sh1=wb["Employe Data"]
row = sh1.max_row
column = sh1.max_column
print(row)
print(column)

for i in range(1,row+1):
    for j in range(1,column+1):
        print(sh1.cell(i,j).value,end="\n")

sh1.cell(row=6,column=1,value="Abhisek")
sh1.cell(row=6,column=2,value="abhisek@test.com")
sh1.cell(row=6,column=3,value="7845898547")
sh1.cell(row=6,column=4,value="105")

wb.save("Emp_demo.xlsx")
