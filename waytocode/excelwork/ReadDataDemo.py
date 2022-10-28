import openpyxl
import xlrd


# workbook = openpyxl.load_workbook(loc)
# workbook = xlrd.open_workbook('Students.xls')
# print(type(workbook))
# sheet = workbook.sheet_by_index(1)
#
# print(sheet.cell_value(1, 1))


# loc = "Employee.xlsx"
# # loc.replace(\,"/")
# print(loc)
# wb = openpyxl.load_workbook(loc)
# sheet = wb.active
#
# row = sheet.max_row
# column = sheet.max_column
# print("total Rows:", row)
# print("total column:", column)
#
# print("\nvalue of first column")
# for i in range(1, row +1):
#     cell= sheet.cell(row=i,column = 1)
#     print(cell.value)
# print("\nvalue of first row")
# for i in range(1, column +1):
#     cell= sheet.cell(row=2,column = i)
#     print(cell.value, end=" ")


import pandas as pd

data = pd.read_excel('Students.xls')
print(data)


